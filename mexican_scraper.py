# -*- coding: utf-8 -*-

import requests
import time
from bs4 import BeautifulSoup
from pprint import pprint

from openpyxl import Workbook
wb = Workbook()
ws = wb.active


headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36"}

questionList = []

a_file = open("medList.txt", "r")

fileobj=open("medList.txt")
it=iter(fileobj)
lines=[]
while True:
    try:
        line=next(it)
        line=line.strip()
        lines.append(line)
    except StopIteration:
        break


def get_float(string):
    f = ''.join(x for x in string if x.isdigit() or x == '.').rstrip('.')
    return float(f)

def getPrices(tag):
    url = f'https://www.farmaciasanpablo.com.mx/search/?text={tag}'
    r = requests.get(url, headers=headers)
    soup = BeautifulSoup(r.text, 'html.parser')
    print(soup.title.text)

    questions = soup.find_all('div', {'class': 'row fsp-grid'})

    if not questions:
        questionList.append('Null')
    
    for item in questions:
        peso = get_float(item.find('p', {'class': 'item-prize'}).text.strip())

        if not peso:
            convert_usd = "Null"

        convert_usd = round(peso * 0.038, 2)

        questionList.append(convert_usd)
        break
    return


for x in range(316):    
    getPrices(lines[x])


with open('peso.txt', 'w+' ) as f:
    for item in questionList:
        f.write(str(item) + '\n')



# def write_headers():
#     headers = ['Name', 'Retail price']
#     for n,cell in enumerate(ws.iter_cols(min_row=1, max_col=2, max_row=1)):
#         cell[0].value = headers[n]

#     wb.save(filename='mexican_retail.xlsx')


# write_headers()

# for n,item in enumerate(questionList, start=2):
#     for n,cell in enumerate(ws.iter_cols(min_row=n, max_col=3, max_row=n)):
#         if n == 0:
#             cell[0].value = item['name']
#         elif n == 1:
#             cell[0].value = item['Retail_price']

# wb.save(filename='mexican_retail.xlsx')



