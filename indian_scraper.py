# -*- coding: utf-8 -*-

import requests
import time
from bs4 import BeautifulSoup
from pprint import pprint

from statistics import median

from openpyxl import Workbook
wb = Workbook()
ws = wb.active


headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36"}

questionList = []

a_file = open("medList.txt", "r")

fileobj=open("medList.txt")
it=iter(fileobj)
lines=[]
while True:
    try:
        line=next(it)
        line=line.strip()
        lines.append(line)
    except StopIteration:
        break


def get_float(string):
    f = ''.join(x for x in string if x.isdigit() or x == '.').rstrip('.')
    return float(f)

def get_int(string):
    return int(''.join(filter(str.isdigit, string)))

def return_median_prices(l):
    arr = []

    for el in l:
        if 'lakh' in el.lower():
            try:
                arr.append(get_float(el) * 100000)
            except Exception as e:
                print(e)
                arr.append(get_int(el) * 100000)
        else:
            try:
                arr.append(get_float(el))
            except Exception as e:
                print(e)
                arr.append(get_int(el))
    print(arr)
    return median(arr)

def get_float(string):
    f = ''.join(x for x in string if x.isdigit() or x == '.').rstrip('.')
    return float(f)

def getPrices(tag):
    url = f'https://dir.indiamart.com/search.mp?ss={tag}&cityid=70469&cq=New+Delhi&mcatid=139578&catid=568&prdsrc=1'
    r = requests.get(url, headers=headers)
    
    soup = BeautifulSoup(r.text, 'html.parser')

    questions = soup.find_all('div', {'class': 'right-group flx1 df fww pr'})

    if not questions:
        questionList.append('Null')
    
    for item in questions:
        with open('html.txt', 'w+') as f:
            f.write(url + '\n')
            f.write(str(item))

        try:
            indian = return_median_prices([x.text for x in item.findAll(attrs={'class': 'prc cp clr3 fwb fs18 prc cp'})])
        except:
            indian = 0

        #ruppee to dollar
        convert_usd = round(indian * 0.013, 2) if indian else "Null"

        questionList.append(convert_usd)

        break
    return

for x in range(316):
    print(lines[x])
    getPrices(lines[x])


# with open('medList.txt', 'r') as f:
#     for line in f:
#         getPrices(line)

with open('found.txt', 'w+' ) as f:
    for item in questionList:
        f.write(str(item) + '\n')


def write_headers():
    headers = ['Retail price']
    for n,cell in enumerate(ws.iter_cols(min_row=1, max_col=1, max_row=1)):
        cell[0].value = headers[n]

    wb.save(filename='indian_retail.xlsx')

write_headers()

for n,item in enumerate(questionList, start=2):
    for cell in ws.iter_cols(min_row=n, max_col=1, max_row=n):
        cell[0].value = item['Retail_price']

wb.save(filename='indian_retail.xlsx')



