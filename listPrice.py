# -*- coding: utf-8 -*-
"""
Created on Wed Mar  9 11:06:38 2022

@author: Rashid
"""
from fileinput import filename
from openpyxl import Workbook
wb = Workbook()
ws = wb.active


price = [
    {'name': 'Gabapentin', 'Retail_price': '$55retail', 'Discount_price': 'The  price after coupon is$11.89.'}, {'name': 'Atorvastatin', 'Retail_price': '$128retail', 'Discount_price': 'The  price after coupon is$21.15.'}, {'name': 'Metformin', 'Retail_price': '$21retail', 'Discount_price': 'The  price after coupon is$13.57.'}, {'name': 'Amlodipine', 'Retail_price': '$52retail', 'Discount_price': 'The  price after coupon is$25.14.'}, {'name': 'Lisinopril', 'Retail_price': '$15retail', 'Discount_price': 'The  price after coupon is$12.96.'}, {'name': 'Levothyroxine', 'Retail_price': '$12retail', 'Discount_price': 'The  price after coupon is$6.83.'}, {'name': 'Losartan', 'Retail_price': '$52retail', 'Discount_price': 'The  price after coupon is$17.16.'}, {'name': 'Metoprolol ER', 'Retail_price': '$41retail', 'Discount_price': 'The  price after coupon is$6.46.'}, {'name': 'Omeprazole', 'Retail_price': '$110retail', 'Discount_price': 'The  price after coupon is$36.51.'}, {'name': 'Metoprolol ER', 'Retail_price': '$41retail', 'Discount_price': 'The  price after coupon is$6.46.'}, {'name': 'Carvedilol', 'Retail_price': '$44retail', 'Discount_price': 'The  price after coupon is$36.44.'}, {'name': 'Hydrocodone / Acetaminophen, Lorcet, Zolvit', 'Retail_price': '$110retail', 'Discount_price': 'The  price after coupon is$30.89.'}, {'name': 'Furosemide', 'Retail_price': '$13retail', 'Discount_price': 'The  price after coupon is$11.72.'}, {'name': 'Lactulose, Generlac', 'Retail_price': '$34retail', 'Discount_price': 'The  price after coupon is$14.55.'}, {'name': 'Simvastatin', 'Retail_price': '$33retail', 'Discount_price': 'The  price after coupon is$5.04.'}, {'name': 'Hydrochlorothiazide', 'Retail_price': '$12retail', 'Discount_price': 'The  price after coupon is$6.64.'}, {'name': 'Gavilyte-G', 'Retail_price': '$24retail', 'Discount_price': 'The  price after coupon is$16.40.'},
    {'name': 'Pantoprazole', 'Retail_price': '$33retail', 'Discount_price': 'The  price after coupon is$10.35.'}, {'name': 'Tamsulosin', 'Retail_price': '$48retail', 'Discount_price': 'The  price after coupon is$6.99.'}, {'name': 'Potassium Chloride', 'Retail_price': '$329retail', 'Discount_price': 'The  price after coupon is$92.95.'}, {'name': 'Eliquis', 'Retail_price': '$609retail', 'Discount_price': 'The  price after coupon is$553.69.'}, {'name': 'Diclofenac Sodium', 'Retail_price': '$48retail', 'Discount_price': 'The  price after coupon is$17.74.'}, {'name': 'Metformin ER (Glucophage XR)', 'Retail_price': '$15retail', 'Discount_price': 'The  price after coupon is$8.31.'}, {'name': 'Peg 3350 and Electrolytes, Trilyte', 'Retail_price': '$42retail', 'Discount_price': 'The  price after coupon is$13.30.'}, {'name': 'Rosuvastatin', 'Retail_price': '$88retail', 'Discount_price': 'The  price after coupon is$11.53.'}, {'name': 'Chlorhexidine, Paroex', 'Retail_price': '$24retail', 'Discount_price': 'The  price after coupon is$7.48.'}, {'name': 'Tramadol', 'Retail_price': '$50retail', 'Discount_price': 'The  price after coupon is$18.67.'}, {'name': 'Trazodone', 'Retail_price': '$16retail', 'Discount_price': 'The  price after coupon is$13.43.'}, {'name': 'Sertraline', 'Retail_price': '$50retail', 'Discount_price': 'The  price after coupon is$19.70.'}, {'name': 'Clopidogrel', 'Retail_price': '$113retail', 'Discount_price': 'The  price after coupon is$34.06.'}
,{'name': 'Pravastatin', 'Retail_price': '$118retail', 'Discount_price': 'The  price after coupon is$16.76.'}, {'name': 'Hydralazine', 'Retail_price': '$42retail', 'Discount_price': 'The  price after coupon is$10.97.'}, {'name': 'Alprazolam', 'Retail_price': '$21retail', 'Discount_price': 'The  price after coupon is$11.70.'}, {'name': 'Allopurinol', 'Retail_price': '$25retail', 'Discount_price': 'The  price after coupon is$4.80.'}, {'name': 'Montelukast', 'Retail_price': '$312retail', 'Discount_price': 'The  price after coupon is$46.30.'}, {'name': 'Oxycodone / Acetaminophen, Prolate, Nalocet', 'Retail_price': '$209retail', 'Discount_price': 'The  price after coupon is$10.06.'}, {'name': 'Duloxetine', 'Retail_price': '$20retail', 'Discount_price': 'The  price after coupon is$7.69.'}, {'name': 'Gavilyte-C', 'Retail_price': '$22retail', 'Discount_price': 'The  price after coupon is$16.40.'}, {'name': 'Atenolol', 'Retail_price': '$13retail', 'Discount_price': 'The  price after coupon is$11.36.'}, {'name': 'Glipizide', 'Retail_price': '$24retail', 'Discount_price': 'The  price after coupon is$14.13.'}, {'name': 'Sodium Chloride', 'Retail_price': '$28retail', 'Discount_price': 'The  price after coupon is$12.66.'}, {'name': 'Famotidine', 'Retail_price': '$39retail', 'Discount_price': 'The  price after coupon is$27.33.'}, {'name': 'Oxycodone', 'Retail_price': '$64retail', 'Discount_price': 'The  price after coupon is$21.39.'}, {'name': 'Triamcinolone, Triderm, Trianex', 'Retail_price': '$6retail', 'Discount_price': 'The  price after coupon is$3.29.'}, {'name': 'Levetiracetam', 'Retail_price': '$149retail', 'Discount_price': 'The  price after coupon is$33.35.'}, {'name': 'Peg 3350 and Electrolytes, Trilyte', 'Retail_price': '$42retail', 'Discount_price': 'The  price after coupon is$13.30.'}, {'name': 'Carbidopa / Levodopa', 'Retail_price': '$53retail', 'Discount_price': 'The  price after coupon is$26.21.'}
,{'name': 'Glimepiride', 'Retail_price': '$53retail', 'Discount_price': 'The  price after coupon is$13.73.'}, {'name': 'Escitalopram', 'Retail_price': '$62retail', 'Discount_price': 'The  price after coupon is$34.20.'}, {'name': 'Lisinopril / HCTZ', 'Retail_price': '$8retail', 'Discount_price': 'The  price after coupon is$7.19.'}, {'name': 'Clonazepam', 'Retail_price': '$24retail', 'Discount_price': 'The  price after coupon is$12.74.'}, {'name': 'Buspirone', 'Retail_price': '$25retail', 'Discount_price': 'The  price after coupon is$21.29.'}, {'name': 'Gavilyte-N', 'Retail_price': '$40retail', 'Discount_price': 'The  price after coupon is$16.32.'}, {'name': 'Meloxicam', 'Retail_price': '$43retail', 'Discount_price': 'The  price after coupon is$28.73.'}, {'name': 'Warfarin, Jantoven', 'Retail_price': '$18retail', 'Discount_price': 'The  price after coupon is$11.20.'}, {'name': 'Prednisone', 'Retail_price': '$12retail', 'Discount_price': 'The  price after coupon is$6.73.'}, {'name': 'Tizanidine', 'Retail_price': '$39retail', 'Discount_price': 'The  price after coupon is$15.96.'}, {'name': 'Fluoxetine', 'Retail_price': '$27retail', 'Discount_price': 'The  price after coupon is$20.22.'}, {'name': 'Citalopram', 'Retail_price': '$20retail', 'Discount_price': 'The  price after coupon is$18.71.'}, {'name': 'Lorazepam', 'Retail_price': '$26retail', 'Discount_price': 'The  price after coupon is$13.90.'}, {'name': 'Quetiapine', 'Retail_price': '$137retail', 'Discount_price': 'The  price after coupon is$32.64.'}, {'name': 'Baclofen', 'Retail_price': '$28retail', 'Discount_price': 'The  price after coupon is$14.43.'}, {'name': 'Ibuprofen', 'Retail_price': '$14retail', 'Discount_price': 'The  price after coupon is$11.54.'}
,{'name': 'Finasteride', 'Retail_price': '$76retail', 'Discount_price': 'The  price after coupon is$22.00.'}, {'name': 'Spironolactone', 'Retail_price': '$56retail', 'Discount_price': 'The  price after coupon is$16.57.'}, {'name': 'Pregabalin', 'Retail_price': '$529retail', 'Discount_price': 'The  price after coupon is$12.50.'}, {'name': 'Clonidine', 'Retail_price': '$21retail', 'Discount_price': 'The  price after coupon is$5.67.'}, {'name': 'Sucralfate', 'Retail_price': '$32retail', 'Discount_price': 'The  price after coupon is$15.64.'}, {'name': 'Donepezil', 'Retail_price': '$49retail', 'Discount_price': 'The  price after coupon is$10.47.'}, {'name': 'Cyclobenzaprine', 'Retail_price': '$18retail', 'Discount_price': 'The  price after coupon is$11.83.'}, {'name': 'Synthroid, Euthyrox, Levoxyl, Levo-T, Unithroid', 'Retail_price': '$12retail', 'Discount_price': 'The  price after coupon is$6.83.'}, {'name': 'Isosorbide Mononitrate ER', 'Retail_price': '$14retail', 'Discount_price': 'The  price after coupon is$13.55.'}, {'name': 'Venlafaxine ER', 'Retail_price': '$120retail', 'Discount_price': 'The  price after coupon is$35.03.'}, {'name': 'Lamotrigine', 'Retail_price': '$61retail', 'Discount_price': 'The  price after coupon is$39.45.'}, {'name': 'Ammonium Lactate', 'Retail_price': '$49retail', 'Discount_price': 'The  price after coupon is$23.16.'}, {'name': 'Topiramate', 'Retail_price': '$49retail', 'Discount_price': 'The  price after coupon is$22.05.'}, {'name': 'Xarelto', 'Retail_price': '$581retail', 'Discount_price': 'The  price after coupon is$538.87.'}, {'name': 'Divalproex', 'Retail_price': '$100retail', 'Discount_price': 'The  price after coupon is$13.61.'}, {'name': 'Losartan / HCTZ', 'Retail_price': '$30retail', 'Discount_price': 'The  price after coupon is$13.25.'}, {'name': 'Ezetimibe', 'Retail_price': '$99retail', 'Discount_price': 'The  price after coupon is$8.26.'}
,{'name': 'Suprep Bowel Prep', 'Retail_price': '$149retail', 'Discount_price': 'The  price after coupon is$110.94.'}, {'name': 'Ropinirole', 'Retail_price': '$52retail', 'Discount_price': 'The  price after coupon is$21.85.'}, {'name': 'Bupropion XL, Budeprion XL', 'Retail_price': '$48retail', 'Discount_price': 'The  price after coupon is$11.79.'}, {'name': 'Mirtazapine', 'Retail_price': '$46retail', 'Discount_price': 'The  price after coupon is$21.63.'}, {'name': 'Lovastatin', 'Retail_price': '$22retail', 'Discount_price': 'The  price after coupon is$12.63.'}, {'name': 'Fluticasone Propionate', 'Retail_price': '$53retail', 'Discount_price': 'The  price after coupon is$21.46.'}, {'name': 'Propranolol', 'Retail_price': '$33retail', 'Discount_price': 'The  price after coupon is$11.64.'}, {'name': 'Vascepa', 'Retail_price': '$377retail', 'Discount_price': 'The  price after coupon is$83.79.'}, {'name': 'Zolpidem', 'Retail_price': '$59retail', 'Discount_price': 'The  price after coupon is$31.99.'}, {'name': 'Glipizide ER', 'Retail_price': '$26retail', 'Discount_price': 'The  price after coupon is$12.35.'}, {'name': 'Nystatin', 'Retail_price': '$26retail', 'Discount_price': 'The  price after coupon is$14.81.'}, {'name': 'Amitriptyline', 'Retail_price': '$15retail', 'Discount_price': 'The  price after coupon is$8.99.'}, {'name': 'Ketoconazole', 'Retail_price': '$107retail', 'Discount_price': 'The  price after coupon is$32.95.'}, {'name': 'Breo Ellipta', 'Retail_price': '$454retail', 'Discount_price': 'The  price after coupon is$380.04.'}, {'name': 'Januvia', 'Retail_price': '$627retail', 'Discount_price': 'The retail price  is$553.86.'}, {'name': 'Celecoxib', 'Retail_price': '$201retail', 'Discount_price': 'The  price after coupon is$12.87.'}, {'name': 'Oxybutynin', 'Retail_price': '$39retail', 'Discount_price': 'The  price after coupon is$17.17.'}, {'name': 'Hydroxyzine Hydrochloride', 'Retail_price': '$29retail', 'Discount_price': 'The  price after coupon is$10.48.'}, {'name': 'Enalapril', 'Retail_price': '$35retail', 'Discount_price': 'The  price after coupon is$30.01.'}, {'name': 'Dicyclomine', 'Retail_price': '$60retail', 'Discount_price': 'The  price after coupon is$22.67.'}, {'name': 'Sodium Chloride', 'Retail_price': '$28retail', 'Discount_price': 'The  price after coupon is$12.66.'}, {'name': 'Nifedipine ER (Procardia XL), Nifedical XL', 'Retail_price': '$41retail', 'Discount_price': 'The  price after coupon is$12.03.'}, {'name': 'Naproxen', 'Retail_price': '$15retail', 'Discount_price': 'The  price after coupon is$11.46.'}, {'name': 'Paroxetine', 'Retail_price': '$35retail', 'Discount_price': 'The  price after coupon is$20.25.'}, {'name': 'Benazepril', 'Retail_price': '$21retail', 'Discount_price': 'The  price after coupon is$13.31.'}, {'name': 'Hydroxychloroquine', 'Retail_price': '$214retail', 'Discount_price': 'The  price after coupon is$13.41.'}, {'name': 'Risperidone', 'Retail_price': '$102retail', 'Discount_price': 'The  price after coupon is$35.85.'}, {'name': 'Bupropion SR, Budeprion SR', 'Retail_price': '$66retail', 'Discount_price': 'The  price after coupon is$7.59.'}, {'name': 'Sevelamer Carbonate', 'Retail_price': '$226retail', 'Discount_price': 'The  price after coupon is$28.66.'}, {'name': 'Pioglitazone', 'Retail_price': '$71retail', 'Discount_price': 'The  price after coupon is$12.38.'}, {'name': 'Torsemide', 'Retail_price': '$24retail', 'Discount_price': 'The  price after coupon is$11.97.'}, {'name': 'Oxybutynin ER', 'Retail_price': '$84retail', 'Discount_price': 'The  price after coupon is$12.28.'}, {'name': 'Advair', 'Retail_price': '$444retail', 'Discount_price': 'The  price after coupon is$89.88.'}, {'name': 'Esomeprazole', 'Retail_price': '$281retail', 'Discount_price': 'The  price after coupon is$11.47.'}, {'name': 'Pramipexole', 'Retail_price': '$79retail', 'Discount_price': 'The  price after coupon is$31.35.'}, {'name': 'Doxazosin', 'Retail_price': '$81retail', 'Discount_price': 'The  price after coupon is$31.03.'}, {'name': 'Enulose', 'Retail_price': '$34retail', 'Discount_price': 'The  price after coupon is$14.55.'}, {'name': 'Valsartan', 'Retail_price': '$71retail', 'Discount_price': 'The  price after coupon is$27.45.'}, {'name': 'Irbesartan', 'Retail_price': '$85retail', 'Discount_price': 'The  price after coupon is$13.94.'}, {'name': 'Ramipril', 'Retail_price': '$49retail', 'Discount_price': 'The  price after coupon is$24.39.'}, {'name': 'Constulose', 'Retail_price': '$34retail', 'Discount_price': 'The  price after coupon is$14.55.'}, {'name': 'Cephalexin', 'Retail_price': '$27retail', 'Discount_price': 'The  price after coupon is$17.67.'}, {'name': 'Diazepam', 'Retail_price': '$15retail', 'Discount_price': 'The  price after coupon is$7.33.'}, {'name': 'Primidone', 'Retail_price': '$40retail', 'Discount_price': 'The  price after coupon is$18.08.'}, {'name': 'Olmesartan', 'Retail_price': '$280retail', 'Discount_price': 'The  price after coupon is$11.75.'}, {'name': 'Omega-3-Acid Ethyl Esters', 'Retail_price': '$211retail', 'Discount_price': 'The  price after coupon is$26.14.'}, {'name': 'Chlorthalidone', 'Retail_price': '$35retail', 'Discount_price': 'The  price after coupon is$5.61.'}, {'name': 'Amoxicillin', 'Retail_price': '$12retail', 'Discount_price': 'The  price after coupon is$8.99.'}, {'name': 'Valproic Acid', 'Retail_price': '$60retail', 'Discount_price': 'The  price after coupon is$24.80.'}, {'name': 'Meclizine', 'Retail_price': '$16retail', 'Discount_price': 'The  price after coupon is$10.13.'}, {'name': 'Labetalol', 'Retail_price': '$120retail', 'Discount_price': 'The  price after coupon is$59.37.'}, {'name': 'Trelegy Ellipta', 'Retail_price': '$725retail', 'Discount_price': 'The  price after coupon is$632.94.'}, {'name': 'Morphine ER', 'Retail_price': '$68retail', 'Discount_price': 'The  price after coupon is$20.44.'}, {'name': 'Acetaminophen / Codeine', 'Retail_price': '$15retail', 'Discount_price': 'The  price after coupon is$10.96.'}, {'name': 'Amiodarone', 'Retail_price': '$54retail', 'Discount_price': 'The  price after coupon is$28.14.'}, {'name': 'Carbamazepine, Epitol', 'Retail_price': '$50retail', 'Discount_price': 'The  price after coupon is$29.22.'}, {'name': 'Terazosin', 'Retail_price': '$19retail', 'Discount_price': 'The  price after coupon is$14.53.'}, {'name': 'Restasis', 'Retail_price': '$688retail', 'Discount_price': 'The  price after coupon is$213.85.'}, {'name': 'Bumetanide', 'Retail_price': '$28retail', 'Discount_price': 'The  price after coupon is$21.08.'}, {'name': 'Estradiol, Yuvafem, Dotti', 'Retail_price': '$80retail', 'Discount_price': 'The  price after coupon is$30.86.'}, {'name': 'Albuterol', 'Retail_price': '$45retail', 'Discount_price': 'The  price after coupon is$26.34.'}, {'name': 'Fenofibrate', 'Retail_price': '$70retail', 'Discount_price': 'The  price after coupon is$6.93.'}, {'name': 'Divalproex ER', 'Retail_price': '$205retail', 'Discount_price': 'The  price after coupon is$13.62.'}, {'name': 'Anoro Ellipta', 'Retail_price': '$531retail', 'Discount_price': 'The  price after coupon is$441.81.'}, {'name': 'Acyclovir', 'Retail_price': '$22retail', 'Discount_price': 'The  price after coupon is$19.19.'}, {'name': 'Creon', 'Retail_price': '$1,169retail', 'Discount_price': 'The  price after coupon is$1,057.28.'}, {'name': 'Sotalol, Sorine', 'Retail_price': '$136retail', 'Discount_price': 'The  price after coupon is$80.21.'}, {'name': 'Jardiance', 'Retail_price': '$582retail', 'Discount_price': 'The  price after coupon is$550.56.'}, {'name': 'Phenytoin', 'Retail_price': '$41retail', 'Discount_price': 'The  price after coupon is$20.35.'}
,{'name': 'Flecainide', 'Retail_price': '$69retail', 'Discount_price': 'The  price after coupon is$20.79.'}, {'name': 'Amlodipine / Benazepril', 'Retail_price': '$80retail', 'Discount_price': 'The  price after coupon is$13.16.'}, {'name': 'Myrbetriq', 'Retail_price': '$517retail', 'Discount_price': 'The  price after coupon is$449.42.'}, {'name': 'Hydrocortisone, Ala-Cort, Proctozone HC', 'Retail_price': '$78retail', 'Discount_price': 'The  price after coupon is$27.13.'}, {'name': 'Olanzapine', 'Retail_price': '$306retail', 'Discount_price': 'The  price after coupon is$80.20.'}, {'name': 'Aripiprazole', 'Retail_price': '$804retail', 'Discount_price': 'The  price after coupon is$12.95.'}, {'name': 'Benztropine Mesylate', 'Retail_price': '$11retail', 'Discount_price': 'The  price after coupon is$8.82.'}, {'name': 'Gemfibrozil', 'Retail_price': '$30retail', 'Discount_price': 'The  price after coupon is$27.80.'}, {'name': 'Metoclopramide', 'Retail_price': '$14retail', 'Discount_price': 'The  price after coupon is$5.45.'}, {'name': 'Oxcarbazepine', 'Retail_price': '$148retail', 'Discount_price': 'The  price after coupon is$38.20.'}, {'name': 'Sodium Chloride', 'Retail_price': '$28retail', 'Discount_price': 'The  price after coupon is$12.66.'}, {'name': 'Doxycycline Hyclate', 'Retail_price': '$76retail', 'Discount_price': 'The  price after coupon is$24.13.'}, {'name': 'Digoxin, Digitek', 'Retail_price': '$20retail', 'Discount_price': 'The  price after coupon is$16.60.'}, {'name': 'Ondansetron', 'Retail_price': '$140retail', 'Discount_price': 'The  price after coupon is$33.25.'}, {'name': 'Calcium Acetate, Eliphos', 'Retail_price': '$134retail', 'Discount_price': 'The  price after coupon is$46.79.'}, {'name': 'Fenofibrate', 'Retail_price': '$70retail', 'Discount_price': 'The  price after coupon is$6.93.'}, {'name': 'Verapamil ER', 'Retail_price': '$27retail', 'Discount_price': 'The  price after coupon is$11.38.'}, {'name': 'Valsartan / HCTZ', 'Retail_price': '$99retail', 'Discount_price': 'The  price after coupon is$19.11.'}, {'name': 'Wixela Inhub, Fluticasone / Salmeterol', 'Retail_price': '$315retail', 'Discount_price': 'The  price after coupon is$124.44.'}, {'name': 'Nortriptyline', 'Retail_price': '$13retail', 'Discount_price': 'The  price after coupon is$11.49.'}, {'name': 'Megestrol', 'Retail_price': '$62retail', 'Discount_price': 'The  price after coupon is$16.58.'}, {'name': 'Promethazine', 'Retail_price': '$21retail', 'Discount_price': 'The  price after coupon is$10.49.'}, {'name': 'Methotrexate', 'Retail_price': '$53retail', 'Discount_price': 'The  price after coupon is$19.18.'}, {'name': 'Midodrine', 'Retail_price': '$229retail', 'Discount_price': 'The  price after coupon is$30.78.'}, {'name': 'Janumet', 'Retail_price': '$645retail', 'Discount_price': 'The  price after coupon is$513.34.'}, {'name': 'Ventolin', 'Retail_price': '$45retail', 'Discount_price': 'The  price after coupon is$28.23.'}, {'name': 'Hydroxyzine Pamoate', 'Retail_price': '$24retail', 'Discount_price': 'The  price after coupon is$11.64.'}
,{'name': 'Golytely', 'Retail_price': '$42retail', 'Discount_price': 'The  price after coupon is$13.30.'}, {'name': 'Pradaxa', 'Retail_price': '$571retail', 'Discount_price': 'The  price after coupon is$478.93.'}, {'name': 'Sulfamethoxazole / Trimethoprim', 'Retail_price': '$16retail', 'Discount_price': 'The  price after coupon is$13.48.'}, {'name': 'Methocarbamol', 'Retail_price': '$20retail', 'Discount_price': 'The  price after coupon is$10.16.'}, {'name': 'Anastrozole', 'Retail_price': '$84retail', 'Discount_price': 'The  price after coupon is$13.27.'}, {'name': 'Valacyclovir', 'Retail_price': '$114retail', 'Discount_price': 'The  price after coupon is$18.55.'}, {'name': 'Bystolic', 'Retail_price': '$172retail', 'Discount_price': 'The  price after coupon is$27.70.'}, {'name': 'Levocetirizine', 'Retail_price': '$80retail', 'Discount_price': 'The  price after coupon is$12.04.'}, {'name': 'Temazepam', 'Retail_price': '$21retail', 'Discount_price': 'The  price after coupon is$13.10.'}, {'name': 'Diltiazem', 'Retail_price': '$22retail', 'Discount_price': 'The  price after coupon is$21.81.'}, {'name': 'Lantus', 'Retail_price': '$353retail', 'Discount_price': 'The  price after coupon is$194.82.'}, {'name': 'Sulfasalazine', 'Retail_price': '$31retail', 'Discount_price': 'The  price after coupon is$15.97.'}, {'name': 'Nitroglycerin, Minitran', 'Retail_price': '$25retail', 'Discount_price': 'The  price after coupon is$10.90.'}, {'name': 'Brilinta', 'Retail_price': '$503retail', 'Discount_price': 'The  price after coupon is$415.62.'}, {'name': 'Cholestyramine, Cholestyramine Light', 'Retail_price': '$57retail', 'Discount_price': 'The  price after coupon is$34.80.'}, {'name': 'Amphetamine Salt Combo', 'Retail_price': '$46retail', 'Discount_price': 'The  price after coupon is$17.74.'}, {'name': 'Cilostazol', 'Retail_price': '$75retail', 'Discount_price': 'The  price after coupon is$27.73.'}, {'name': 'Clobetasol, Embeline', 'Retail_price': '$321retail', 'Discount_price': 'The  price after coupon is$16.56.'}, {'name': 'Methylprednisolone', 'Retail_price': '$20retail', 'Discount_price': 'The  price after coupon is$11.18.'}, {'name': 'Klor-Con M', 'Retail_price': '$21retail', 'Discount_price': 'The  price after coupon is$12.29.'}, {'name': 'Mesalamine', 'Retail_price': '$1,255retail', 'Discount_price': 'The  price after coupon is$122.54.'}, {'name': 'Tradjenta', 'Retail_price': '$589retail', 'Discount_price': 'The  price after coupon is$436.99.'}, {'name': 'Methadone', 'Retail_price': '$36retail', 'Discount_price': 'The  price after coupon is$16.53.'}
,{'name': 'Lidocaine', 'Retail_price': '$268retail', 'Discount_price': 'The  price after coupon is$62.20.'}, {'name': 'Euthyrox, Synthroid, Unithroid, Levo-T, Levoxyl', 'Retail_price': '$12retail', 'Discount_price': 'The  price after coupon is$6.83.'}, {'name': 'Lansoprazole', 'Retail_price': '$169retail', 'Discount_price': 'The  price after coupon is$33.53.'}, {'name': 'Mycophenolate', 'Retail_price': '$793retail', 'Discount_price': 'The  price after coupon is$31.33.'}, {'name': 'Quinapril', 'Retail_price': '$64retail', 'Discount_price': 'The  price after coupon is$22.75.'}, {'name': 'Linzess', 'Retail_price': '$584retail', 'Discount_price': 'The  price after coupon is$514.08.'}, {'name': 'Ipratropium / Albuterol', 'Retail_price': '$53retail', 'Discount_price': 'The  price after coupon is$17.80.'}, {'name': 'Hydromorphone', 'Retail_price': '$67retail', 'Discount_price': 'The  price after coupon is$10.42.'}, {'name': 'Carbidopa / Levodopa ER', 'Retail_price': '$130retail', 'Discount_price': 'The  price after coupon is$37.30.'}, {'name': 'Ciprofloxacin', 'Retail_price': '$23retail', 'Discount_price': 'The  price after coupon is$21.39.'}, {'name': 'Alendronate', 'Retail_price': '$42retail', 'Discount_price': 'The  price after coupon is$12.21.'}, {'name': 'Generlac, Lactulose', 'Retail_price': '$32retail', 'Discount_price': 'The  price after coupon is$14.55.'}, {'name': 'Clozapine', 'Retail_price': '$276retail', 'Discount_price': 'The  price after coupon is$64.19.'}, {'name': 'Ondansetron', 'Retail_price': '$140retail', 'Discount_price': 'The  price after coupon is$33.25.'}, {'name': 'Telmisartan', 'Retail_price': '$224retail', 'Discount_price': 'The  price after coupon is$13.41.'}, {'name': 'Azelastine', 'Retail_price': '$59retail', 'Discount_price': 'The  price after coupon is$25.83.'}, {'name': 'Mupirocin', 'Retail_price': '$53retail', 'Discount_price': 'The  price after coupon is$19.14.'}, {'name': 'Venlafaxine', 'Retail_price': '$57retail', 'Discount_price': 'The  price after coupon is$20.02.'}, {'name': 'Peg 3350 and Electrolytes, Trilyte', 'Retail_price': '$42retail', 'Discount_price': 'The  price after coupon is$13.30.'}, {'name': 'Vimpat', 'Retail_price': '$1,211retail', 'Discount_price': 'The  price after coupon is$1,095.50.'}
,{'name': 'Lidocaine', 'Retail_price': '$268retail', 'Discount_price': 'The  price after coupon is$62.20.'}, {'name': 'Euthyrox, Synthroid, Unithroid, Levo-T, Levoxyl', 'Retail_price': '$12retail', 'Discount_price': 'The  price after coupon is$6.83.'}, {'name': 'Lansoprazole', 'Retail_price': '$169retail', 'Discount_price': 'The  price after coupon is$33.53.'}, {'name': 'Mycophenolate', 'Retail_price': '$793retail', 'Discount_price': 'The  price after coupon is$31.33.'}, {'name': 'Quinapril', 'Retail_price': '$64retail', 'Discount_price': 'The  price after coupon is$22.75.'}, {'name': 'Linzess', 'Retail_price': '$584retail', 'Discount_price': 'The  price after coupon is$514.08.'}, {'name': 'Ipratropium / Albuterol', 'Retail_price': '$53retail', 'Discount_price': 'The  price after coupon is$17.80.'}, {'name': 'Hydromorphone', 'Retail_price': '$67retail', 'Discount_price': 'The  price after coupon is$10.42.'}, {'name': 'Carbidopa / Levodopa ER', 'Retail_price': '$130retail', 'Discount_price': 'The  price after coupon is$37.30.'}, {'name': 'Ciprofloxacin', 'Retail_price': '$23retail', 'Discount_price': 'The  price after coupon is$21.39.'}, {'name': 'Alendronate', 'Retail_price': '$42retail', 'Discount_price': 'The  price after coupon is$12.21.'}, {'name': 'Generlac, Lactulose', 'Retail_price': '$40retail', 'Discount_price': 'The  price after coupon is$14.55.'}, {'name': 'Clozapine', 'Retail_price': '$276retail', 'Discount_price': 'The  price after coupon is$64.19.'}, {'name': 'Ondansetron', 'Retail_price': '$140retail', 'Discount_price': 'The  price after coupon is$33.25.'}, {'name': 'Telmisartan', 'Retail_price': '$224retail', 'Discount_price': 'The  price after coupon is$13.41.'}, {'name': 'Azelastine', 'Retail_price': '$59retail', 'Discount_price': 'The  price after coupon is$25.83.'}, {'name': 'Mupirocin', 'Retail_price': '$53retail', 'Discount_price': 'The  price after coupon is$19.14.'}, {'name': 'Venlafaxine', 'Retail_price': '$57retail', 'Discount_price': 'The  price after coupon is$20.02.'}, {'name': 'Peg 3350 and Electrolytes, Trilyte', 'Retail_price': '$42retail', 'Discount_price': 'The  price after coupon is$13.30.'}, {'name': 'Vimpat', 'Retail_price': '$1,211retail', 'Discount_price': 'The  price after coupon is$1,095.50.'}, {'name': 'Doxepin', 'Retail_price': '$19retail', 'Discount_price': 'The  price after coupon is$8.45.'}
,{'name': 'Potassium Citrate ER', 'Retail_price': '$84retail', 'Discount_price': 'The  price after coupon is$17.09.'}, {'name': 'Dexilant', 'Retail_price': '$299retail', 'Discount_price': 'The  price after coupon is$166.63.'}, {'name': 'Cartia XT, Diltiazem ER (Cardizem CD), Dilt-Cd', 'Retail_price': '$95retail', 'Discount_price': 'The  price after coupon is$20.45.'}, {'name': 'Metronidazole', 'Retail_price': '$19retail', 'Discount_price': 'The  price after coupon is$10.30.'}, {'name': 'Lithium Carbonate', 'Retail_price': '$12retail', 'Discount_price': 'The  price after coupon is$11.25.'}, {'name': 'Propranolol ER', 'Retail_price': '$56retail', 'Discount_price': 'The  price after coupon is$8.46.'}, {'name': 'Calcitriol', 'Retail_price': '$43retail', 'Discount_price': 'The  price after coupon is$9.69.'}, {'name': 'Clotrimazole / Betamethasone', 'Retail_price': '$69retail', 'Discount_price': 'The  price after coupon is$15.04.'}, {'name': 'Testosterone', 'Retail_price': '$666retail', 'Discount_price': 'The  price after coupon is$51.80.'}, {'name': 'Lidocaine', 'Retail_price': '$268retail', 'Discount_price': 'The  price after coupon is$62.20.'}, {'name': 'Prazosin', 'Retail_price': '$26retail', 'Discount_price': 'The  price after coupon is$11.33.'}, {'name': 'Dexamethasone, Dexpak, Hidex', 'Retail_price': '$20retail', 'Discount_price': 'The  price after coupon is$11.50.'}, {'name': 'Isosorbide Dinitrate', 'Retail_price': '$86retail', 'Discount_price': 'The  price after coupon is$28.06.'}, {'name': 'Bisoprolol', 'Retail_price': '$49retail', 'Discount_price': 'The  price after coupon is$21.85.'}, {'name': 'Clindamycin', 'Retail_price': '$105retail', 'Discount_price': 'The  price after coupon is$28.09.'}, {'name': 'Phenobarbital', 'Retail_price': '$42retail', 'Discount_price': 'The  price after coupon is$17.64.'}, {'name': 'Diphenoxylate / Atropine', 'Retail_price': '$55retail', 'Discount_price': 'The  price after coupon is$21.02.'}, {'name': 'Carbamazepine ER', 'Retail_price': '$109retail', 'Discount_price': 'The  price after coupon is$36.73.'}, {'name': 'Fluticasone / Salmeterol, Wixela Inhub', 'Retail_price': '$444retail', 'Discount_price': 'The  price after coupon is$89.88.'}, {'name': 'Raloxifene', 'Retail_price': '$469retail', 'Discount_price': 'The  price after coupon is$40.99.'}, {'name': 'Repaglinide', 'Retail_price': '$677retail', 'Discount_price': 'The  price after coupon is$83.63.'}, {'name': 'Symbicort', 'Retail_price': '$281retail', 'Discount_price': 'The  price after coupon is$210.83.'}, {'name': 'Morphine', 'Retail_price': '$55retail', 'Discount_price': 'The  price after coupon is$25.48.'}, {'name': 'Tacrolimus, Hecoria', 'Retail_price': '$260retail', 'Discount_price': 'The  price after coupon is$37.45.'}, {'name': 'Bupropion', 'Retail_price': '$25retail', 'Discount_price': 'The  price after coupon is$6.67.'}, {'name': 'Liothyronine', 'Retail_price': '$79retail', 'Discount_price': 'The  price after coupon is$29.17.'}, {'name': 'Nabumetone', 'Retail_price': '$72retail', 'Discount_price': 'The  price after coupon is$20.16.'}, {'name': 'Carafate', 'Retail_price': '$32retail', 'Discount_price': 'The  price after coupon is$15.64.'}, {'name': 'Bisoprolol / HCTZ', 'Retail_price': '$41retail', 'Discount_price': 'The  price after coupon is$34.63.'}, {'name': 'Olmesartan / HCTZ', 'Retail_price': '$290retail', 'Discount_price': 'The  price after coupon is$17.16.'}, {'name': 'Ursodiol', 'Retail_price': '$80retail', 'Discount_price': 'The  price after coupon is$19.12.'}, {'name': 'Zonisamide', 'Retail_price': '$263retail', 'Discount_price': 'The  price after coupon is$15.84.'}, {'name': 'Dutasteride', 'Retail_price': '$250retail', 'Discount_price': 'The  price after coupon is$14.03.'}, {'name': 'Glipizide / Metformin', 'Retail_price': '$171retail', 'Discount_price': 'The  price after coupon is$50.49.'}, {'name': 'Alfuzosin', 'Retail_price': '$108retail', 'Discount_price': 'The  price after coupon is$14.29.'}, {'name': 'Oxycontin', 'Retail_price': '$280retail', 'Discount_price': 'The  price after coupon is$51.98.'}, {'name': 'Incruse Ellipta', 'Retail_price': '$417retail', 'Discount_price': 'The  price after coupon is$360.16.'}, {'name': 'Welchol', 'Retail_price': '$225retail', 'Discount_price': 'The  price after coupon is$31.96.'}, {'name': 'Levemir', 'Retail_price': '$490retail', 'Discount_price': 'The  price after coupon is$462.36.'}, {'name': 'Amantadine', 'Retail_price': '$61retail', 'Discount_price': 'The  price after coupon is$21.73.'}, {'name': 'Azithromycin', 'Retail_price': '$40retail', 'Discount_price': 'The  price after coupon is$17.57.'}, {'name': 'Haloperidol', 'Retail_price': '$22retail', 'Discount_price': 'The  price after coupon is$15.38.'}, {'name': 'Butalbital / Acetaminophen / Caffeine', 'Retail_price': '$43retail', 'Discount_price': 'The  price after coupon is$18.41.'}, {'name': 'Dilt-Xr, Diltiazem ER (Dilacor XR)', 'Retail_price': '$44retail', 'Discount_price': 'The  price after coupon is$21.95.'}, {'name': 'Colestipol', 'Retail_price': '$53retail', 'Discount_price': 'The  price after coupon is$29.55.'}, {'name': 'Novolog, Relion', 'Retail_price': '$174retail', 'Discount_price': 'The  price after coupon is$58.83.'}, {'name': 'Azathioprine', 'Retail_price': '$94retail', 'Discount_price': 'The  price after coupon is$19.24.'}, {'name': 'Methimazole', 'Retail_price': '$21retail', 'Discount_price': 'The  price after coupon is$11.95.'}, {'name': 'Nitrofurantoin Mono/Macro', 'Retail_price': '$55retail', 'Discount_price': 'The  price after coupon is$16.31.'}, {'name': 'Buprenorphine / Naloxone', 'Retail_price': '$83retail', 'Discount_price': 'The  price after coupon is$43.55.'}, {'name': 'Rytary', 'Retail_price': '$389retail', 'Discount_price': 'The  price after coupon is$363.52.'}, {'name': 'Vancomycin', 'Retail_price': '$1,039retail', 'Discount_price': 'The  price after coupon is$78.10.'}, {'name': 'Letrozole', 'Retail_price': '$130retail', 'Discount_price': 'The  price after coupon is$8.67.'}, {'name': 'Renvela', 'Retail_price': '$226retail', 'Discount_price': 'The  price after coupon is$28.66.'}, {'name': 'Loperamide', 'Retail_price': '$21retail', 'Discount_price': 'The  price after coupon is$14.86.'}, {'name': 'Methylphenidate', 'Retail_price': '$92retail', 'Discount_price': 'The  price after coupon is$29.35.'}, {'name': 'Irbesartan / HCTZ', 'Retail_price': '$103retail', 'Discount_price': 'The  price after coupon is$29.77.'}, {'name': 'Jantoven, Warfarin', 'Retail_price': '$51retail', 'Discount_price': 'The  price after coupon is$14.20.'}, {'name': 'Fluocinonide', 'Retail_price': '$104retail', 'Discount_price': 'The  price after coupon is$17.45.'}, {'name': 'Ipratropium', 'Retail_price': '$85retail', 'Discount_price': 'The  price after coupon is$15.62.'}, {'name': 'Amitiza', 'Retail_price': '$308retail', 'Discount_price': 'The  price after coupon is$157.69.'}, {'name': 'Farxiga', 'Retail_price': '$629retail', 'Discount_price': 'The  price after coupon is$529.74.'}
]

def write_headers():
    headers = ['Name', 'Retail price', 'Discount price']
    for n,cell in enumerate(ws.iter_cols(min_row=1, max_col=3, max_row=1)):
        cell[0].value = headers[n]

    wb.save(filename='drugrx.xlsx')

write_headers()

# return only integers from string
def get_int(string):
    return int(''.join(filter(str.isdigit, string)))

# return only float from string
def get_float(string):
    f = ''.join(x for x in string if x.isdigit() or x == '.').rstrip('.')
    return float(f)


for n,item in enumerate(price, start=2):
    for n,cell in enumerate(ws.iter_cols(min_row=n, max_col=3, max_row=n)):
        if n == 0:
            cell[0].value = item['name']
        elif n == 1:
            cell[0].value = get_int(item['Retail_price'])
        else:
            cell[0].value = get_float(item['Discount_price'])
        

wb.save(filename='drugrx.xlsx')
